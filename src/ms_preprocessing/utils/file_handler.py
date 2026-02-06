"""
File handling utilities for MS Preprocessing Toolkit.

This module provides functions for reading and writing various file formats
commonly used in mass spectrometry data processing.
"""

from pathlib import Path
from typing import Optional, Union, Tuple
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color


class FileHandler:
    """
    Handles file I/O operations for mass spectrometry data files.

    Supports Excel (.xlsx, .xls), CSV, and TSV formats with
    preservation of formatting information where applicable.
    """

    SUPPORTED_FORMATS = {".xlsx", ".xls", ".csv", ".tsv", ".txt"}

    def __init__(self):
        """Initialize the FileHandler."""
        self._last_loaded_path: Optional[Path] = None
        self._red_font_rows: set = set()

    @staticmethod
    def is_supported_format(file_path: Union[str, Path]) -> bool:
        """Check if the file format is supported."""
        path = Path(file_path)
        return path.suffix.lower() in FileHandler.SUPPORTED_FORMATS

    def load_data(
        self,
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = 0,
        header_row: int = 0,
    ) -> Tuple[pd.DataFrame, dict]:
        """
        Load data from a file.

        Args:
            file_path: Path to the input file
            sheet_name: Sheet name or index for Excel files
            header_row: Row index to use as header

        Returns:
            Tuple of (DataFrame, metadata dict)
        """
        path = Path(file_path)
        self._last_loaded_path = path
        self._red_font_rows = set()
        metadata = {"source_file": str(path), "load_time": datetime.now().isoformat()}

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if not self.is_supported_format(path):
            raise ValueError(f"Unsupported file format: {path.suffix}")

        suffix = path.suffix.lower()

        if suffix in {".xlsx", ".xls"}:
            df = self._load_excel(path, sheet_name, header_row)
            metadata["format"] = "excel"
            metadata["sheet_name"] = sheet_name
        elif suffix == ".csv":
            df = pd.read_csv(path, header=header_row)
            metadata["format"] = "csv"
        elif suffix in {".tsv", ".txt"}:
            df = pd.read_csv(path, sep="\t", header=header_row)
            metadata["format"] = "tsv"
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

        metadata["shape"] = df.shape
        metadata["columns"] = list(df.columns)
        metadata["red_font_rows"] = sorted(self._red_font_rows)

        return df, metadata

    def _load_excel(
        self,
        file_path: Path,
        sheet_name: Optional[Union[str, int]] = 0,
        header_row: int = 0,
    ) -> pd.DataFrame:
        """Load data from Excel file with formatting extraction."""
        # Load with pandas
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

        # Extract red font information for protection logic
        self._red_font_rows = set()
        try:
            wb = load_workbook(file_path, data_only=False)
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb[sheet_name] if sheet_name else wb.active

            # Check each row for red font (starting from data rows)
            for row_idx in range(header_row + 2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == "rgb" and color.rgb:
                        # Check for red color (various shades)
                        rgb = color.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            r = int(rgb[-6:-4], 16)
                            g = int(rgb[-4:-2], 16)
                            b = int(rgb[-2:], 16)
                            if r > 200 and g < 100 and b < 100:
                                # Adjust for pandas DataFrame index (0-based, excluding header)
                                self._red_font_rows.add(row_idx - header_row - 2)
            wb.close()
        except Exception:
            # If formatting extraction fails, continue without it
            pass

        return df

    def get_red_font_rows(self) -> set:
        """Get the set of row indices with red font."""
        return self._red_font_rows

    def save_data(
        self,
        df: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        index: bool = False,
        highlight_rows: Optional[set] = None,
        blue_font_cells: Optional[list] = None,
        red_font_rows: Optional[set] = None,
        extra_sheets: Optional[dict] = None,
    ) -> Path:
        """
        Save data to a file.

        Args:
            df: DataFrame to save
            file_path: Output file path
            sheet_name: Sheet name for Excel files
            index: Whether to include DataFrame index
            highlight_rows: Set of row indices to highlight with yellow
            blue_font_cells: List of (row, col) tuples for blue font cells

        Returns:
            Path to the saved file
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix in {".xlsx", ".xls"}:
            self._save_excel(
                df,
                path,
                sheet_name,
                index,
                highlight_rows,
                blue_font_cells,
                red_font_rows,
                extra_sheets=extra_sheets,
            )
        elif suffix == ".csv":
            df.to_csv(path, index=index)
        elif suffix in {".tsv", ".txt"}:
            df.to_csv(path, sep="\t", index=index)
        else:
            # Default to Excel format
            path = path.with_suffix(".xlsx")
            self._save_excel(df, path, sheet_name, index, highlight_rows, blue_font_cells, red_font_rows)

        return path

    def _save_excel(
        self,
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str = "Sheet1",
        index: bool = False,
        highlight_rows: Optional[set] = None,
        blue_font_cells: Optional[list] = None,
        red_font_rows: Optional[set] = None,
        extra_sheets: Optional[dict] = None,
    ) -> None:
        """Save DataFrame to Excel with optional formatting."""
        # First save with pandas
        if extra_sheets:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                for sheet, sheet_df in extra_sheets.items():
                    if sheet_df is None:
                        continue
                    sheet_df.to_excel(writer, sheet_name=sheet, index=index)
        else:
            df.to_excel(file_path, sheet_name=sheet_name, index=index)

        # Then apply formatting if needed
        if highlight_rows or blue_font_cells or red_font_rows:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            blue_font = Font(color="0070C0")
            red_font = Font(color="FF0000")

            # Apply yellow highlight to specified rows
            if highlight_rows:
                for row_idx in highlight_rows:
                    # Account for header row (Excel is 1-indexed, header is row 1)
                    excel_row = row_idx + 2
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=excel_row, column=col).fill = yellow_fill

            # Apply blue font to specified cells
            if blue_font_cells:
                for row_idx, col_idx in blue_font_cells:
                    excel_row = row_idx + 2
                    excel_col = col_idx + 1
                    ws.cell(row=excel_row, column=excel_col).font = blue_font

            # Apply red font to specified rows (FeatureID column)
            if red_font_rows:
                for row_idx in red_font_rows:
                    excel_row = row_idx + 2
                    ws.cell(row=excel_row, column=1).font = red_font

            wb.save(file_path)
            wb.close()

    @staticmethod
    def generate_output_path(
        input_path: Union[str, Path],
        suffix: str = "_processed",
        output_dir: Optional[Union[str, Path]] = None,
        add_timestamp: bool = True,
    ) -> Path:
        """
        Generate an output file path based on input path.

        Args:
            input_path: Original input file path
            suffix: Suffix to add before file extension
            output_dir: Output directory (defaults to input file directory)
            add_timestamp: Whether to add timestamp to filename

        Returns:
            Generated output path
        """
        input_path = Path(input_path)
        stem = input_path.stem

        if add_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_name = f"{stem}{suffix}_{timestamp}{input_path.suffix}"
        else:
            new_name = f"{stem}{suffix}{input_path.suffix}"

        if output_dir:
            output_path = Path(output_dir) / new_name
        else:
            output_path = input_path.parent / new_name

        return output_path


def parse_mz_rt_string(value: str) -> Tuple[Optional[float], Optional[float]]:
    """
    Parse a combined m/z and RT string (e.g., "123.456/1.23").

    Args:
        value: String in format "mz/rt"

    Returns:
        Tuple of (mz, rt) as floats, or (None, None) if parsing fails
    """
    try:
        parts = str(value).split("/")
        if len(parts) == 2:
            mz = float(parts[0].strip())
            rt = float(parts[1].strip())
            return mz, rt
    except (ValueError, AttributeError):
        pass
    return None, None


def format_mz_rt_string(mz: float, rt: float, mz_decimals: int = 4, rt_decimals: int = 2) -> str:
    """
    Format m/z and RT values as a combined string.

    Args:
        mz: m/z value
        rt: RT value
        mz_decimals: Decimal places for m/z
        rt_decimals: Decimal places for RT

    Returns:
        Formatted string in "mz/rt" format
    """
    return f"{mz:.{mz_decimals}f}/{rt:.{rt_decimals}f}"
