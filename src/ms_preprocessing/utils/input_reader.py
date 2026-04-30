"""Input readers for toolkit-supported data formats."""

from __future__ import annotations

import importlib.util
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class InputReader:
    """Read Excel and delimited data while extracting input metadata."""

    @staticmethod
    def load_excel(
        file_path: Path,
        sheet_name: str | int | None = 0,
        header_row: int = 0,
    ) -> tuple[pd.DataFrame, set[int]]:
        """Load an Excel sheet and return red-font row indexes."""
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, engine="openpyxl")

        red_font_rows: set[int] = set()
        try:
            wb = load_workbook(file_path, data_only=False)
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb[sheet_name] if sheet_name else wb.active

            for row_idx in range(header_row + 2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=1)
                if cell.font and cell.font.color:
                    color = cell.font.color
                    if color.type == "rgb" and color.rgb:
                        rgb = color.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            r = int(rgb[-6:-4], 16)
                            g = int(rgb[-4:-2], 16)
                            b = int(rgb[-2:], 16)
                            if r > 200 and g < 100 and b < 100:
                                red_font_rows.add(row_idx - header_row - 2)
            wb.close()
        except Exception as exc:
            logger.warning("Failed to extract red-font formatting: %s", exc)

        return df, red_font_rows

    @staticmethod
    def load_delimited(path: Path, header_row: int, sep: str) -> pd.DataFrame:
        """Load CSV/TSV using pyarrow engine if available."""
        engine = "pyarrow" if importlib.util.find_spec("pyarrow") is not None else None

        if engine:
            try:
                return pd.read_csv(path, header=header_row, sep=sep, engine=engine)
            except Exception as exc:
                logger.debug(
                    "pyarrow delimited parse failed for %s; falling back to default parser: %s",
                    path,
                    exc,
                )
        return pd.read_csv(path, header=header_row, sep=sep)
