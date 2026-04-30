"""Excel writing and formatting helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class ExcelFormattingWriter:
    """Write Excel workbooks and apply row/cell formatting metadata."""

    @staticmethod
    def save(
        df: pd.DataFrame,
        file_path: Path,
        sheet_name: str = "Sheet1",
        index: bool = False,
        highlight_rows: set | None = None,
        blue_font_cells: list | None = None,
        red_font_rows: set | None = None,
        extra_sheets: dict | None = None,
    ) -> None:
        if extra_sheets:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
                for sheet, sheet_df in extra_sheets.items():
                    if sheet_df is None:
                        continue
                    sheet_df.to_excel(writer, sheet_name=sheet, index=index)
        else:
            df.to_excel(file_path, sheet_name=sheet_name, index=index)

        if not (highlight_rows or blue_font_cells or red_font_rows):
            return

        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_font = Font(color="0070C0")
        red_font = Font(color="FF0000")

        if highlight_rows:
            for row_idx in highlight_rows:
                excel_row = row_idx + 2
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=excel_row, column=col).fill = yellow_fill

        if blue_font_cells:
            for row_idx, col_idx in blue_font_cells:
                excel_row = row_idx + 2
                excel_col = col_idx + 1
                ws.cell(row=excel_row, column=excel_col).font = blue_font

        if red_font_rows:
            for row_idx in red_font_rows:
                excel_row = row_idx + 2
                ws.cell(row=excel_row, column=1).font = red_font

        wb.save(file_path)
        wb.close()
